"""ממיר את מאגר הכתובות הפתוח ל-CSV מוכן לייבוא לטבלת addresses.

מקור: "רשימת כתובות בישראל עם קואורדינטות", odata.org.il, רישיון CC-BY.
הקואורדינטות במקור הן רשת ישראל החדשה (EPSG:2039) ומומרות כאן ל-WGS84.

שני קמטים במקור שהסקריפט מיישר:
  1. כ-30% מהשורות ארציות מגיעות עם X=0,Y=0 — בלי קואורדינטה. מסוננות.
  2. עמודת מספר הבית אינה אחידה: לרוב "20", אבל לרחובות ששמם במקור
     בסדר הפוך ("רגר יצחק") היא מכילה את הכתובת המלאה בסדר טבעי
     ("יצחק רגר 154"). מנצלים את זה — הטקסט הזה הוא שם הרחוב כפי
     שאדם מקליד אותו, אז הוא מועדף לתצוגה.

שימוש:
    python scripts/build_addresses_csv.py <input.xlsx> <output.csv> [עיר]

בלי ארגומנט עיר — מייצא את כל הארץ.
"""

import csv
import re
import sys

import openpyxl
from pyproj import Transformer

HOUSE_NUMBER = re.compile(r"(\d+\s*[א-ת]?)\s*$")


def main() -> int:
    if len(sys.argv) < 3:
        print(__doc__)
        return 1

    src, dest = sys.argv[1], sys.argv[2]
    only_city = sys.argv[3] if len(sys.argv) > 3 else None

    to_wgs84 = Transformer.from_crs("EPSG:2039", "EPSG:4326", always_xy=True)

    workbook = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    next(rows)  # כותרת

    kept = skipped_no_coords = skipped_no_number = 0

    with open(dest, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["city", "street", "house_number", "neighbourhood", "lat", "lng"])

        for city, street, number, neighbourhood, _district, x, y in rows:
            if only_city and city != only_city:
                continue
            if not x or not y:
                skipped_no_coords += 1
                continue

            raw = str(number or "").strip()
            match = HOUSE_NUMBER.search(raw)
            if not match:
                skipped_no_number += 1
                continue

            house = match.group(1).replace(" ", "")
            natural_name = raw[: match.start()].strip()
            display_street = natural_name or (street or "").strip()
            if not display_street:
                skipped_no_number += 1
                continue

            lng, lat = to_wgs84.transform(x, y)
            writer.writerow(
                [city, display_street, house, (neighbourhood or "").strip() or None, f"{lat:.6f}", f"{lng:.6f}"]
            )
            kept += 1

    print(f"נכתבו {kept} שורות ל-{dest}")
    print(f"דולגו: {skipped_no_coords} בלי קואורדינטות, {skipped_no_number} בלי מספר בית")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
